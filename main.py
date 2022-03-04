##########################################################
# Author: John Luong
# Email: lvminh97@gmail.com
##########################################################

from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook as WB
from time import sleep

__DRIVERPATH__      = "driver/chromedriver.exe"
__URL__             = "https://dichvucong.quangninh.gov.vn/Default.aspx?tabid=119";
__EXPORTFILE__      = "export/dichvucong.xlsx"

browser = webdriver.Chrome(executable_path = __DRIVERPATH__)
browser.get(__URL__)
# Wait 5 seconds for full loading of website
sleep(5)

data = []

# Crawl data from all pages
for page in range(79):
    for i in range(10):
        try:
            row = browser.find_element(by = By.ID, value = "dnn_ctr507_HoiDap_dgDanhSachHoSo_ctl00__" + str(i))
            cols = row.find_elements(by = By.TAG_NAME, value = "td")
            temp = {}
            temp['id'] = cols[0].text
            tmpStr = cols[1].find_element(by = By.TAG_NAME, value = "a").text
            tmpStr = tmpStr.strip()
            tmpStr.replace("&nbsp;", "")
            temp['cauhoi'] = tmpStr
            tmpStr = cols[2].text
            tmpStr = tmpStr.strip()
            tmpStr.replace("&nbsp;", "")
            temp['linhvuc'] = tmpStr
            data += [temp]
        except:
            continue
    
    print("Finish %d/79 pages" % (page + 1))        
    browser.find_element(by = By.CLASS_NAME, value = "rgPageNext").click()
    sleep(1)

# Export data to Excel file
output = WB()
sheet = output.active
for i in range(len(data)):
    sheet.cell(row = i + 1, column = 1).value = data[i]['id']
    sheet.cell(row = i + 1, column = 2).value = data[i]['cauhoi']
    sheet.cell(row = i + 1, column = 3).value = data[i]['linhvuc']
output.save(__EXPORTFILE__)

# Close browser
browser.quit()
